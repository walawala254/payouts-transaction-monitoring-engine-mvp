from copy import deepcopy
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from transaction_monitor import (
    ALERT_COLUMNS,
    build_alert_cases,
    effective_limits,
    evaluate,
    load_rules,
    make_excel_report,
    normalize_transactions,
    spreadsheet_safe_dataframe,
    simulate_rule_change,
    validate_upload,
)


ROOT = Path(__file__).resolve().parents[1]


def configured_rules():
    return load_rules(ROOT / "rules.yaml")


def raw_transactions(rows):
    return normalize_transactions(pd.DataFrame(rows))


def base_row(**changes):
    row = {
        "transaction id": "tx-1",
        "merchant": "Merchant A",
        "mid": "MID-A",
        "transaction date": "2026-01-01 10:00:00",
        "transaction type": "payment",
        "status": "approved",
        "amount": 100,
        "merchant amount": 95,
        "currency": "USD",
        "card brand": "visa",
        "card number": "4111111111111111",
        "payer email": "person@example.com",
        "payer phone": "+254 700 000 001",
        "payer country": "KE",
        "geoip country": "KE",
    }
    row.update(changes)
    return row


def test_every_monitoring_rule_has_governance_schema():
    rules = configured_rules()
    required = {
        "code", "name", "category", "classification", "purpose", "risk_scenario_detected",
        "applicable_entity", "required_data_fields", "data_quality_dependencies",
        "eligibility_conditions", "exclusions", "baseline_method", "threshold_logic",
        "time_window", "aggregation_key", "risk_weight", "severity",
        "applicable_antifraud_plans", "merchant_applicability", "mid_applicability",
        "override_behavior", "alert_message", "evidence_returned",
        "recommended_analyst_action", "false_positive_risks", "false_positive_controls",
        "rule_dependencies", "testing_status", "enabled",
    }
    assert rules["rule_catalog"]
    for rule in rules["rule_catalog"]:
        assert rule["classification"] == "Monitoring Rule"
        assert required <= set(rule)


def test_preventive_limits_are_context_not_tm_alerts():
    tx = raw_transactions([base_row(amount=6000, **{"merchant amount": 6000})])
    evaluated, alerts = evaluate(tx, configured_rules())
    assert "HIGH_AMOUNT_ABOVE_PLAN_LIMIT" in evaluated.loc[0, "preventive_control_breaches"]
    assert "HIGH_AMOUNT_ABOVE_PLAN_LIMIT" not in set(alerts["rule"])
    assert "MERCHANT_VALUE_ABOVE_LIMIT" not in set(alerts["rule"])


def test_specific_and_temporary_inheritance_wins():
    rules = configured_rules()
    rules["merchant_overrides"] = {"Merchant A": {"max_amount": 800}}
    rules["mid_overrides"] = {"MID-A": {"max_amount": 700}}
    rules["temporary_overrides"] = [{
        "id": "TEMP-1", "merchant": "Merchant A", "mid": "MID-A", "enabled": True,
        "starts_at": "2026-01-01", "expires_at": "2026-01-31",
        "values": {"max_amount": 900},
    }]
    row = raw_transactions([base_row()]).iloc[0]
    limits = effective_limits(row, rules)
    assert limits["max_amount"] == 900
    assert limits["rule_scope"] == "temporary:TEMP-1"


def test_velocity_cases_and_duplicate_transaction_ids_use_row_identity():
    rules = configured_rules()
    rules["plans"]["STANDARD"]["card_approved_count_limit"] = 1
    rows = [
        base_row(**{"transaction id": "duplicate", "transaction date": "2026-01-01 10:00:00"}),
        base_row(**{"transaction id": "duplicate", "transaction date": "2026-01-01 10:05:00"}),
    ]
    evaluated, alerts = evaluate(raw_transactions(rows), rules)
    assert len(evaluated) == 2
    assert evaluated["risk_score"].gt(0).sum() == 1
    assert "card_number" not in ALERT_COLUMNS
    assert "card_display" in alerts
    cases = build_alert_cases(evaluated, alerts)
    assert not cases.empty
    assert cases.iloc[0]["triggered_rule_count"] >= 1


def test_trailing_baseline_spike_activates_only_after_history():
    rules = configured_rules()
    rules["baseline_settings"]["minimum_history_days"] = 3
    rules["global"]["merchant_volume_baseline_multiplier"] = 3
    rows = []
    for day in range(1, 8):
        rows.append(base_row(**{
            "transaction id": f"tx-{day}",
            "transaction date": f"2026-01-{day:02d} 10:00:00",
            "amount": 1000 if day == 7 else 100,
            "card number": f"411111111111{day:04d}",
        }))
    _, alerts = evaluate(raw_transactions(rows), rules)
    spike_hits = alerts[alerts["rule"] == "MERCHANT_VOLUME_BASELINE_SPIKE"]
    assert len(spike_hits) == 1
    assert float(spike_hits.iloc[0]["baseline_value"]) == 100


def test_simulation_returns_current_proposed_comparison():
    current = configured_rules()
    proposed = deepcopy(current)
    proposed["plans"]["STANDARD"]["card_approved_count_limit"] = 1
    tx = raw_transactions([
        base_row(**{"transaction id": "tx-1", "transaction date": "2026-01-01 10:00:00"}),
        base_row(**{"transaction id": "tx-2", "transaction date": "2026-01-01 10:05:00"}),
    ])
    comparison, affected = simulate_rule_change(tx, current, proposed)
    assert comparison["setting"].tolist() == ["Current", "Proposed"]
    assert comparison.iloc[1]["transactions_matched"] >= comparison.iloc[0]["transactions_matched"]
    assert not affected.empty


def test_enriched_excel_preserves_cases_hits_and_raw_transactions():
    rules = configured_rules()
    rules["plans"]["STANDARD"]["card_approved_count_limit"] = 1
    tx = raw_transactions([
        base_row(**{"transaction id": "tx-1", "transaction date": "2026-01-01 10:00:00"}),
        base_row(**{"transaction id": "tx-2", "transaction date": "2026-01-01 10:05:00"}),
    ])
    evaluated, hits = evaluate(tx, rules)
    cases = build_alert_cases(evaluated, hits)
    report = make_excel_report(evaluated, hits, cases, rules)
    workbook = load_workbook(BytesIO(report), read_only=True)
    assert {"Investigation Alerts", "Rule Hits", "Normalized Transactions", "Data Quality"} <= set(workbook.sheetnames)


def test_public_configuration_and_sample_are_synthetic():
    rules = configured_rules()
    assert rules["configuration_metadata"]["profile"] == "PUBLIC_DEMO"
    assert rules["configuration_metadata"]["synthetic"] is True
    assert all(str(mid).startswith("DEMO_") for mid in rules["mid_rules"])
    assert not any(rules.get("watchlists", {}).values())

    sample = pd.read_csv(ROOT / "sample_data" / "synthetic_transactions.csv")
    validation = validate_upload(sample, max_rows=100_000)
    assert validation[validation["severity"] == "Error"].empty
    evaluated, hits = evaluate(normalize_transactions(sample), rules)
    assert len(evaluated) == len(sample)
    assert not hits.empty


def test_upload_validation_blocks_missing_required_fields_and_row_limit():
    invalid = pd.DataFrame({"transaction id": ["tx-1", "tx-2"]})
    issues = validate_upload(invalid, max_rows=1)
    assert "Row limit exceeded" in set(issues["issue"])
    assert "Required mapping missing" in set(issues["issue"])
    assert set(issues["severity"]) == {"Error"}


def test_spreadsheet_exports_neutralize_formula_like_text():
    source = pd.DataFrame({"value": ["=1+1", "+SUM(A1:A2)", "-2+3", "@cmd", "normal", 10]})
    safe = spreadsheet_safe_dataframe(source)
    assert safe.loc[:3, "value"].str.startswith("'").all()
    assert safe.loc[4, "value"] == "normal"
    assert safe.loc[5, "value"] == 10
